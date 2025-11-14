from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.http import HttpResponse, HttpResponseForbidden
import csv
import datetime

from reportlab.lib.pagesizes import A4

from apps.reserva.models import Reserva
from apps.recorrido.models import Recorrido, Parada
from django.db.models import Count, Sum
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# VISTA DEL DASHBOARD

@login_required
def dashboard_informes(request):
    if request.user.rol.nombre != 'ADMINISTRADOR':
        return HttpResponseForbidden("No tienes permiso para acceder a esta página.")

    # Obtenemos todos los recorridos para pasarlos al template
    # y que el admin pueda elegir uno en el formulario.
    recorridos_disponibles = Recorrido.objects.filter(estado='activo')
    # ... (código sin cambios) ...

    contexto = {
        'recorridos': recorridos_disponibles
    }
    return render(request, 'informe/dashboard.html', contexto)


# VISTA DE PASAJEROS POR FECHA
@login_required
def generar_informe_pasajeros_fechas(request):
    if request.user.rol.nombre != 'ADMINISTRADOR':
        return HttpResponseForbidden("Acceso denegado")

    # 2. Obtener datos de la URL
    fecha_inicio_str = request.GET.get('fecha_inicio', None)
    fecha_fin_str = request.GET.get('fecha_fin', None)
    formato = request.GET.get('formato', 'pdf')

    if not (fecha_inicio_str and fecha_fin_str):
        return redirect('dashboard_informes')

    try:
        fecha_inicio = datetime.datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
        fecha_fin = datetime.datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
    except ValueError:
        return HttpResponse("Fechas inválidas. Use el formato AAAA-MM-DD.")

    if fecha_inicio > fecha_fin:
        return HttpResponse(
            "Error: La 'Fecha de Inicio' no puede ser posterior a la 'Fecha de Fin'.",
            status=400
        )

    # 3. Consulta a la BD
    reservas = Reserva.objects.filter(
        itinerario__fecha_itinerario__range=[fecha_inicio, fecha_fin],
        estado='C'
    ).select_related(
        'turista',
        'itinerario__recorrido'
    ).order_by('itinerario__fecha_itinerario')

    # 4. Calcular el total
    # ... (código sin cambios) ...
    total_pasajeros = 0
    for reserva in reservas:
        total_pasajeros += reserva.cantidad_asientos

    # --- LOGICA PARA CSV ---
    if formato == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="reporte_pasajeros_{fecha_inicio_str}_al_{fecha_fin_str}.csv"'
        writer = csv.writer(response)
        writer.writerow(['ID Reserva', 'Fecha Viaje', 'Recorrido', 'Nombre Turista', 'Asientos'])

        for reserva in reservas:
            writer.writerow([
                reserva.id,
                reserva.itinerario.fecha_itinerario,
                reserva.itinerario.recorrido.nombre_recorrido,
                reserva.turista.get_full_name(),
                reserva.cantidad_asientos
            ])

        writer.writerow([])
        writer.writerow(['Total de Pasajeros:', total_pasajeros])
        return response

    # --- LOGICA PARA PDF ---
    elif formato == 'pdf':
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="reporte_pasajeros_{fecha_inicio_str}_al_{fecha_fin_str}.pdf"'

        p = canvas.Canvas(response, pagesize=A4)
        width, height = A4
        y = height - (2 * cm)
        x = 2 * cm
        p.setFont("Helvetica-Bold", 16)
        # ... (código sin cambios) ...
        p.drawString(x, y, "Reporte de Pasajeros por Fecha")
        y -= 0.5 * cm
        p.line(x, y, width - (2 * cm), y);
        y -= 1 * cm
        p.setFont("Helvetica", 12)
        # ... (código sin cambios) ...
        p.drawString(x, y, f"Período: {fecha_inicio_str} al {fecha_fin_str}")
        y -= 0.5 * cm
        p.drawString(x, y, "Estado: CONFIRMADAS");
        y -= 1 * cm
        p.setFont("Helvetica-Bold", 11)
        p.drawString(x, y, "Fecha Viaje")
        p.drawString(x + (4 * cm), y, "Recorrido")
        p.drawString(x + (10 * cm), y, "Nombre Turista")
        p.drawString(x + (14 * cm), y, "Asientos")
        y -= 0.5 * cm
        p.setFont("Helvetica", 10)

        for reserva in reservas:
            p.drawString(x, y, str(reserva.itinerario.fecha_itinerario))
            p.drawString(x + (4 * cm), y, reserva.itinerario.recorrido.nombre_recorrido[:30])
            p.drawString(x + (10 * cm), y, reserva.turista.get_full_name()[:25])
            p.drawString(x + (14 * cm), y, str(reserva.cantidad_asientos))
            y -= 0.7 * cm
            if y < (3 * cm): y = height - (2 * cm); p.showPage(); p.setFont("Helvetica", 10)

        y -= 0.5 * cm
        p.line(x, y, width - (2 * cm), y)
        y -= 0.5 * cm
        p.setFont("Helvetica-Bold", 12)
        p.drawString(x, y, f"Total de Pasajeros: {total_pasajeros}")

        p.showPage()
        p.save()
        return response

    # --- LÓGICA PARA EXCEL ---
    elif formato == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="reporte_pasajeros_{fecha_inicio_str}_al_{fecha_fin_str}.xlsx"'
        wb = Workbook()
        ws = wb.active
        ws.title = "Pasajeros por Fecha"

        header_font = Font(bold=True, name='Arial', size=12)
        center_align = Alignment(horizontal='center')

        ws['A1'] = "Reporte de Pasajeros por Fecha"
        ws['A1'].font = Font(bold=True, name='Arial', size=16)
        ws['A2'] = f"Período: {fecha_inicio_str} al {fecha_fin_str}"
        ws['A3'] = "Estado: CONFIRMADAS"

        headers = ['ID Reserva', 'Fecha Viaje', 'Recorrido', 'Nombre Turista', 'Asientos']
        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_num, value=header_title)
            cell.font = header_font
            ws.column_dimensions[chr(64 + col_num)].width = 25

        row_num = 6
        for reserva in reservas:
            ws.append([
                reserva.id,
                reserva.itinerario.fecha_itinerario,
                reserva.itinerario.recorrido.nombre_recorrido,
                reserva.turista.get_full_name(),
                reserva.cantidad_asientos
            ])
            ws.cell(row=row_num, column=5).alignment = center_align
            row_num += 1

        total_row = row_num + 1
        ws.cell(row=total_row, column=4, value="Total de Pasajeros:").font = header_font
        ws.cell(row=total_row, column=5, value=total_pasajeros).font = header_font
        ws.cell(row=total_row, column=5).alignment = center_align

        wb.save(response)
        return response

    return redirect('dashboard_informes')

# --- VISTA DE RESERVAS POR RECORRIDO ---
@login_required
def generar_informe_por_recorrido(request):
    """
    Genera un reporte de reservas para un recorrido específico,
    en un rango de fechas.
    """
    # 1. Seguridad
    if request.user.rol.nombre != 'ADMINISTRADOR':
        return HttpResponseForbidden("Acceso denegado")

    # 2. Obtener datos de la URL
    fecha_inicio_str = request.GET.get('fecha_inicio', None)
    fecha_fin_str = request.GET.get('fecha_fin', None)
    formato = request.GET.get('formato', 'pdf')
    recorrido_id = request.GET.get('recorrido_id', None)

    # 3. Validar datos
    if not (fecha_inicio_str and fecha_fin_str and recorrido_id):
        return redirect('dashboard_informes')

    try:
        fecha_inicio = datetime.datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
        fecha_fin = datetime.datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
        recorrido = Recorrido.objects.get(pk=recorrido_id)  # Verificamos que el recorrido exista
    except (ValueError, Recorrido.DoesNotExist):
        return HttpResponse("Datos inválidos (Fecha o Recorrido no existe).", status=400)

    if fecha_inicio > fecha_fin:
        # ... (código sin cambios) ...
        return HttpResponse(
            "Error: La 'Fecha de Inicio' no puede ser posterior a la 'Fecha de Fin'.",
            status=400
        )

    # 4. Consulta a la BD
    reservas = Reserva.objects.filter(
        itinerario__fecha_itinerario__range=[fecha_inicio, fecha_fin],
        itinerario__recorrido_id=recorrido_id,  # <-- ¡EL NUEVO FILTRO!
        estado='C'
    ).select_related(
        'turista',
        'itinerario__recorrido'
    ).order_by('itinerario__fecha_itinerario')

    # 5. Calcular el total
    total_pasajeros = 0
    for reserva in reservas:
        total_pasajeros += reserva.cantidad_asientos

    titulo_reporte = f"Reservas para '{recorrido.nombre_recorrido}'"
    subtitulo_periodo = f"Período: {fecha_inicio_str} al {fecha_fin_str}"

    if formato == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="reporte_por_recorrido_{fecha_inicio_str}_al_{fecha_fin_str}.csv"'
        writer = csv.writer(response)
        writer.writerow([titulo_reporte])
        writer.writerow([subtitulo_periodo])
        writer.writerow(['ID Reserva', 'Fecha Viaje', 'Nombre Turista', 'Asientos'])

        for reserva in reservas:
            writer.writerow([
                reserva.id,
                reserva.itinerario.fecha_itinerario,
                reserva.turista.get_full_name(),
                reserva.cantidad_asientos
            ])

        writer.writerow([])
        writer.writerow(['Total de Pasajeros:', total_pasajeros])
        return response

    # --- LOGICA PARA PDF ---
    elif formato == 'pdf':
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="reporte_por_recorrido_{fecha_inicio_str}_al_{fecha_fin_str}.pdf"'
        p = canvas.Canvas(response, pagesize=A4)
        width, height = A4
        y = height - (2 * cm)
        x = 2 * cm
        # ... (código sin cambios) ...
        p.setFont("Helvetica-Bold", 16)
        p.drawString(x, y, titulo_reporte)
        y -= 1 * cm
        p.setFont("Helvetica", 12)
        p.drawString(x, y, subtitulo_periodo)
        y -= 0.5 * cm
        p.drawString(x, y, "Estado: CONFIRMADAS")
        y -= 1 * cm
        p.setFont("Helvetica-Bold", 11)
        p.drawString(x, y, "Fecha Viaje")
        p.drawString(x + (6 * cm), y, "Nombre Turista")
        p.drawString(x + (12 * cm), y, "Asientos")
        y -= 0.5 * cm
        p.setFont("Helvetica", 10)

        for reserva in reservas:
            p.drawString(x, y, str(reserva.itinerario.fecha_itinerario))
            p.drawString(x + (6 * cm), y, reserva.turista.get_full_name()[:30])
            p.drawString(x + (12 * cm), y, str(reserva.cantidad_asientos))
            y -= 0.7 * cm
            if y < (3 * cm): y = height - (2 * cm); p.showPage(); p.setFont("Helvetica", 10)

        y -= 0.5 * cm
        p.line(x, y, width - (2 * cm), y)
        y -= 0.5 * cm
        p.setFont("Helvetica-Bold", 12)
        p.drawString(x, y, f"Total de Pasajeros: {total_pasajeros}")
        p.showPage()
        p.save()
        return response

    # --- LÓGICA PARA EXCEL ---
    elif formato == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="reporte_por_recorrido_{fecha_inicio_str}_al_{fecha_fin_str}.xlsx"'
        wb = Workbook()
        ws = wb.active
        ws.title = "Reservas por Recorrido"
        header_font = Font(bold=True, name='Arial', size=12)
        center_align = Alignment(horizontal='center')

        ws['A1'] = titulo_reporte
        ws['A1'].font = Font(bold=True, name='Arial', size=16)
        ws['A2'] = subtitulo_periodo
        ws['A3'] = "Estado: CONFIRMADAS"
        headers = ['ID Reserva', 'Fecha Viaje', 'Nombre Turista', 'Asientos']
        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_num, value=header_title)
            cell.font = header_font;
            ws.column_dimensions[chr(64 + col_num)].width = 25

        row_num = 6
        for reserva in reservas:
            ws.append([
                reserva.id,
                reserva.itinerario.fecha_itinerario,
                reserva.turista.get_full_name(),
                reserva.cantidad_asientos
            ])
            ws.cell(row=row_num, column=4).alignment = center_align
            row_num += 1

        total_row = row_num + 1
        ws.cell(row=total_row, column=3, value="Total de Pasajeros:").font = header_font
        ws.cell(row=total_row, column=4, value=total_pasajeros).font = header_font
        ws.cell(row=total_row, column=4).alignment = center_align
        wb.save(response)
        return response

    return redirect('dashboard_informes')

# --- VISTA PARA RECORRIDOS ACTIVOS ---
@login_required()
def generar_informe_recorridos_activos(request):
    """
    Genera un listado simple de todos los recorridos
    cuyo estado es 'activo'. No usa fechas.
    """
    # 1. Seguridad
    if request.user.rol.nombre != 'ADMINISTRADOR':
        return HttpResponseForbidden("Acceso denegado")

    # 2. Obtener formato (no hay fechas)
    formato = request.GET.get('formato', 'pdf')

    # 3. Consulta a la BD
    recorridos = Recorrido.objects.filter(estado='activo').order_by('nombre_recorrido')

    # Obtener solo la fecha actual
    hoy = datetime.date.today()
    # Formatear como string (DD/MM/AAAA)
    fecha_actual_str = hoy.strftime("%d/%m/%Y")

    titulo_reporte = f"Listado de Recorridos Activos | Fecha: {fecha_actual_str}"

    # --- LOGICA PARA CSV ---
    if formato == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="recorridos_activos_{fecha_actual_str}.csv"'
        writer = csv.writer(response)
        writer.writerow([titulo_reporte])
        writer.writerow(['ID', 'Nombre', 'Descripción', 'Duración (min)', 'Precio'])

        for r in recorridos:
            writer.writerow([
                r.id,
                r.nombre_recorrido,
                r.descripcion_recorrido,
                r.duracion,
                r.precio
            ])
        return response

    # --- LOGICA PARA PDF ---
    elif formato == 'pdf':
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="recorridos_activos_{fecha_actual_str}.pdf"'
        p = canvas.Canvas(response, pagesize=A4)
        width, height = A4
        y = height - (2 * cm)
        x = 2 * cm
        p.setFont("Helvetica-Bold", 16)
        p.drawString(x, y, titulo_reporte)
        y -= 1 * cm
        p.setFont("Helvetica-Bold", 11)
        p.drawString(x, y, "Nombre")
        p.drawString(x + (5 * cm), y, "Duración (min)")
        p.drawString(x + (9 * cm), y, "Precio")
        y -= 0.5 * cm
        p.setFont("Helvetica", 10)

        for r in recorridos:
            p.drawString(x, y, r.nombre_recorrido[:30])
            p.drawString(x + (5 * cm), y, str(r.duracion))
            p.drawString(x + (9 * cm), y, f"${r.precio}")
            y -= 0.7 * cm
            if y < (3 * cm): y = height - (2 * cm); p.showPage(); p.setFont("Helvetica", 10)

        p.showPage()
        p.save()
        return response

    # --- LOGICA PARA EXCEL ---
    elif formato == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="recorridos_activos_{fecha_actual_str}.xlsx"'
        wb = Workbook()
        ws = wb.active
        ws.title = "Recorridos Activos"
        header_font = Font(bold=True, name='Arial', size=12)

        ws['A1'] = titulo_reporte
        ws['A1'].font = Font(bold=True, name='Arial', size=16)
        headers = ['ID', 'Nombre', 'Descripción', 'Duración (min)', 'Precio']
        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num, value=header_title)
            cell.font = header_font
            ws.column_dimensions[chr(64 + col_num)].width = 25

        for r in recorridos:
            ws.append([
                r.id,
                r.nombre_recorrido,
                r.descripcion_recorrido,
                r.duracion,
                r.precio
            ])
        wb.save(response)
        return response

    return redirect('dashboard_informes')

# --- VISTA PARA PARADAS MÁS UTILIZADAS ---
@login_required
def generar_informe_paradas_utilizadas(request):
    """
    Genera un reporte de paradas (como punto_partida)
    agrupadas por cantidad de reservas y pasajeros,
    en un rango de fechas.
    """
    # 1. Seguridad
    if request.user.rol.nombre != 'ADMINISTRADOR':
        return HttpResponseForbidden("Acceso denegado")

    # 2. Obtener datos de la URL
    fecha_inicio_str = request.GET.get('fecha_inicio', None)
    fecha_fin_str = request.GET.get('fecha_fin', None)
    formato = request.GET.get('formato', 'pdf')

    if not (fecha_inicio_str and fecha_fin_str):
        return redirect('dashboard_informes')

    try:
        fecha_inicio = datetime.datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
        fecha_fin = datetime.datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
    except ValueError:
        return HttpResponse("Fechas inválidas. Use el formato AAAA-MM-DD.")

    if fecha_inicio > fecha_fin:
        return HttpResponse(
            "Error: La 'Fecha de Inicio' no puede ser posterior a la 'Fecha de Fin'.",
            status=400
        )

    # Usamos .values() para agrupar por el nombre de la parada
    # Usamos .annotate() para crear los campos calculados (total_reservas, total_pasajeros)
    # Contamos las reservas confirmadas ('C') en el rango de fechas.
    datos_paradas = Reserva.objects.filter(
        itinerario__fecha_itinerario__range=[fecha_inicio, fecha_fin],
        estado='C'
    ).values(
        'punto_partida__nombre'  # Agrupar por el nombre de la parada
    ).annotate(
        total_reservas=Count('id'),  # Contar cuántas reservas
        total_pasajeros=Sum('cantidad_asientos')  # Sumar cuántos asientos
    ).order_by(
        '-total_pasajeros'  # Ordenar de más pasajeros a menos
    )

    titulo_reporte = "Paradas más Utilizadas (como Punto de Partida)"
    subtitulo_periodo = f"Período: {fecha_inicio_str} al {fecha_fin_str}"

    # --- LOGICA PARA CSV ---
    if formato == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="paradas_utilizadas_{fecha_inicio_str}_al_{fecha_fin_str}.csv"'
        writer = csv.writer(response)
        writer.writerow([titulo_reporte])
        writer.writerow([subtitulo_periodo])
        writer.writerow(['Nombre de la Parada', 'Total Reservas', 'Total Pasajeros'])

        for parada in datos_paradas:
            writer.writerow([
                parada['punto_partida__nombre'],
                parada['total_reservas'],
                parada['total_pasajeros']
            ])
        return response

    # --- LOGICA PARA PDF ---
    elif formato == 'pdf':
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="paradas_utilizadas_{fecha_inicio_str}_al_{fecha_fin_str}.pdf"'
        p = canvas.Canvas(response, pagesize=A4)
        width, height = A4
        y = height - (2 * cm)
        x = 2 * cm
        p.setFont("Helvetica-Bold", 16)
        p.drawString(x, y, titulo_reporte)
        y -= 1 * cm
        p.setFont("Helvetica", 12)
        p.drawString(x, y, subtitulo_periodo)
        y -= 1 * cm
        p.setFont("Helvetica-Bold", 11)
        p.drawString(x, y, "Nombre Parada");
        p.drawString(x + (8 * cm), y, "Total Reservas")
        p.drawString(x + (12 * cm), y, "Total Pasajeros")
        y -= 0.5 * cm
        p.setFont("Helvetica", 10)

        for parada in datos_paradas:
            p.drawString(x, y, parada['punto_partida__nombre'])
            p.drawString(x + (8 * cm), y, str(parada['total_reservas']))
            p.drawString(x + (12 * cm), y, str(parada['total_pasajeros']))
            y -= 0.7 * cm
            if y < (3 * cm): y = height - (2 * cm); p.showPage(); p.setFont("Helvetica", 10)

        p.showPage()
        p.save()
        return response

    # --- LOGICA PARA EXCEL ---
    elif formato == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="paradas_utilizadas_{fecha_inicio_str}_al_{fecha_fin_str}.xlsx"'
        wb = Workbook()
        ws = wb.active
        ws.title = "Paradas Utilizadas"
        header_font = Font(bold=True, name='Arial', size=12)
        center_align = Alignment(horizontal='center')

        ws['A1'] = titulo_reporte;
        ws['A1'].font = Font(bold=True, name='Arial', size=16)
        ws['A2'] = subtitulo_periodo
        headers = ['Nombre Parada', 'Total Reservas', 'Total Pasajeros']
        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num, value=header_title)
            cell.font = header_font;
            ws.column_dimensions[chr(64 + col_num)].width = 30

        row_num = 5
        for parada in datos_paradas:
            ws.cell(row=row_num, column=1, value=parada['punto_partida__nombre'])
            cell_reservas = ws.cell(row=row_num, column=2, value=parada['total_reservas'])
            cell_pasajeros = ws.cell(row=row_num, column=3, value=parada['total_pasajeros'])
            cell_reservas.alignment = center_align
            cell_pasajeros.alignment = center_align
            row_num += 1

        wb.save(response)
        return response

    return redirect('dashboard_informes')